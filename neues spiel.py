import openpyxl as xl
import random
import os
from funktionen import random_race, random_npc_cityloc_even, npc_disappears, create_players
from os.path import expanduser

home = expanduser("~")
wb = xl.load_workbook("BR New Game Blank.xlsx")
day = 1
player_lists = create_players(int(input("Wie viele Spieler sollen mitspielen? ")))
player_names = player_lists[0]
player_locs = player_lists[1]

while day <= 13:
    sheet = wb["Tabelle1"]
    genlist = []
    traders_count = 0
    misc_count = 0
    dissappeared_count = 0
    # Personengruppen werden gezählt
    for row in range(2, sheet.max_row + 1):
        gen_check = sheet.cell(row, 5)
        genlist.append(gen_check.value)
        traders_count = genlist.count("Ja stadt 1")
        misc_count = genlist.count("Ja stadt")

    # Listen werden erstellt
    traders_locs = random_npc_cityloc_even(traders_count)
    misc_locs = random_npc_cityloc_even(misc_count)

    # NPC Einträge in Excel Dateien
    for row in range(2, sheet.max_row + 1):
        gen_check = sheet.cell(row, 5)
        race_entry = sheet.cell(row, 2)
        loc_entry = sheet.cell(row, 3)
        if gen_check.value != "Nein":
            if day == 1:
                race_entry.value = random_race()
        if gen_check.value == "Ja stadt 1" and day == 1:
            random_number = random.randint(1, traders_count) - 1
            loc_entry.value = traders_locs[random_number]
            traders_locs.pop(random_number)
            traders_count -= 1
        if gen_check.value == "Ja stadt":
            random_number = random.randint(1, misc_count) - 1
            loc_entry.value = misc_locs[random_number]
            misc_locs.pop(random_number)
            misc_count -= 1
        if gen_check.value == "Ja wohn+kip":
            loc_entry.value = npc_disappears()

    # PC Einträge
    race_list = []

    for player_name in player_names:
        name_entry = sheet.cell(sheet.max_row + 1, 1)
        name_entry.value = player_names[player_names.index(player_name)]
        if day == 1:
            race = random_race()
            race_list.append(race)
        race_entry = sheet.cell(sheet.max_row, 2)
        print(race_list)
        print(player_names.index(player_name))
        # index error
        race_entry.value = race_list[player_names.index(player_name)]
        loc_entry = sheet.cell(sheet.max_row, 3)
        loc_entry.value = player_locs[player_names.index(player_name)]

    day_entry = sheet.cell(sheet.max_row + 2, 1)
    day_entry.value = f"Tag {day}"
    wb.save(f"BR Tag {day}.xlsx")
    sheet.delete_rows(sheet.max_row - (1 + len(player_names)), sheet.max_row)
    day += 2

day = 1
# Bereinigung der Tabelle
while day <= 13:
    wb = xl.load_workbook(f"BR Tag {day}.xlsx")
    sheet = wb["Tabelle1"]
    sheet.delete_cols(5)
    wb.save(f"{home}/Desktop/BR Tag {day}.xlsx")
    os.remove(f"BR Tag {day}.xlsx")
    day += 2
