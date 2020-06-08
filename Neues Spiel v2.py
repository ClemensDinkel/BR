import openpyxl as xl
import random
import os
from funktionen import random_race, random_npc_cityloc_even, npc_disappears, random_pc_livingloc_even
from listen import locations
from os.path import expanduser


home = expanduser("~")
wb = xl.load_workbook("BR New Game Blank.xlsx")
day = 1
genlist = []
traders_count = 0
misc_count = 0
dissappeared_count = 0

# Spieler Input
player_count = int(input("Wie viele Spieler sollen mitspielen? "))
player_list = []
spielernummer = 1
while spielernummer <= player_count:
    player_list.append(input(f"Spieler {spielernummer}: "))
    spielernummer += 1

# Spieler werden in Liste eingetragen, gen Eintrag
sheet = wb["Tabelle1"]
for player in player_list:
    name_entry = sheet.cell(sheet.max_row + 1, 1)
    func_entry = sheet.cell(sheet.max_row, 4)
    gen_entry = sheet.cell(sheet.max_row, 5)
    name_entry.value = player_list[player_list.index(player)]
    func_entry.value = "Spieler"
    gen_entry.value = "Ja wohn"
# NPC Personengruppen werden gezählt
for row in range(4, sheet.max_row + 1):
    gen_check = sheet.cell(row, 5)
    genlist.append(gen_check.value)
    traders_count = genlist.count("Ja stadt 1")
    misc_count = genlist.count("Ja stadt")
    dissappeared_count = genlist.count("Ja wohn+kip")

# while day nicht vergessen, hier?
# Loc_Listen werden erstellt
traders_locs = random_npc_cityloc_even(traders_count)
misc_locs = random_npc_cityloc_even(misc_count)
pc_locs = random_pc_livingloc_even(player_count)

# Liste mit Dictionaries (Chars) wird erstellt
char_list = []
name = ""
for row in range(4, sheet.max_row + 1):
    name_entry = sheet.cell(row, 1)
    race_entry = sheet.cell(row, 2)
    loc_entry = sheet.cell(row, 3)
    gen_check = sheet.cell(row, 5)
    name = name_entry.value
    if name is not None:
        name = name.lower()
        name = name.replace(" ", "_")
    name = {"name":f"{name_entry}"}
    print(name.get("name"))
# ...

wb.save(f"{home}/Desktop/BR Tag {day}.xlsx")

# Einträge in Excel
