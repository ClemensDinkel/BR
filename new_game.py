import openpyxl as xl
import random
from funktionen import random_id, random_npc_cityloc_even, npc_disappears, random_pc_livingloc_even
from os.path import expanduser

home = expanduser("~")
wb = xl.load_workbook("BR New Game Blank.xlsx")
sheet = wb["Tabelle1"]

day = 1
traders_count = 0
misc_count = 0
dissappeared_count = 0
genlist = []

# Spieler Input, Eintrag in Excel
player_count = int(input("Wie viele Spieler sollen mitspielen? "))
spielernummer = 1
while spielernummer <= player_count:
    name_entry = sheet.cell(sheet.max_row + 1, 1)
    function_entry = sheet.cell(sheet.max_row, 4)
    gen_entry = sheet.cell(sheet.max_row, 5)
    name_entry.value = input(f"Spieler {spielernummer}: ")
    function_entry.value = "Spieler"
    gen_entry.value = "Ja wohn 1"
    spielernummer += 1

while day <= 13:
    sheet = wb["Tabelle1"]
    day_entry = sheet.cell(1, 1)
    day_entry.value = f"Tag {day}"

    # Personengruppen werden gezählt. Wdh nötig, da counts bei Zuweisung von Locations auf 0 gesetzt werden.
    for row2 in range(4, sheet.max_row + 1):
        gen_check = sheet.cell(row2, 5)
        genlist.append(gen_check.value)
        traders_count = genlist.count("Ja stadt 1")
        misc_count = genlist.count("Ja stadt")
        player_count = genlist.count("Ja wohn 1")

    # Loc_Listen werden erstellt
    traders_locs = random_npc_cityloc_even(traders_count)
    misc_locs = random_npc_cityloc_even(misc_count)
    pc_locs = random_pc_livingloc_even(player_count)

    # Excel wird ausgefüllt
    for row in range(4, sheet.max_row + 1):
        id_entry = sheet.cell(row, 2)
        loc_entry = sheet.cell(row, 3)
        if genlist[row - 4] != "Nein" and day == 1:
            id_entry.value = random_id()

        if genlist[row - 4] == "Ja stadt 1" and day == 1:
            random_number = random.randint(1, traders_count) - 1
            loc_entry.value = traders_locs[random_number]
            traders_locs.pop(random_number)
            traders_count -= 1
        if genlist[row - 4] == "Ja wohn 1" and day == 1:
            random_number = random.randint(1, player_count) - 1
            loc_entry.value = pc_locs[random_number]
            pc_locs.pop(random_number)
            player_count -= 1
        if genlist[row - 4] == "Ja stadt":
            random_number = random.randint(1, misc_count) - 1
            loc_entry.value = misc_locs[random_number]
            misc_locs.pop(random_number)
            misc_count -= 1
        if genlist[row - 4] == "Ja wohn+kip":
            loc_entry.value = npc_disappears()

    # letzte Modifikationen für die Ausgabe
    sheet.delete_cols(5)
    wb.save(f"{home}/Desktop/BR Tag {day}.xlsx")
    day += 2
